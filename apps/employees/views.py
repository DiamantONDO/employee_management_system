from rest_framework import status
from rest_framework.viewsets import ViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.core.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
import io

from apps.employees.services import EmployeeService
from apps.employees.serializers import (
    EmployeeSerializer,
    EmployeeCreateSerializer,
    EmployeeUpdateSerializer
)
from apps.employees.filters import EmployeeFilter


class EmployeeViewSet(ViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = EmployeeFilter

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.service = EmployeeService()

    def _get_filtered_employees(self, request):
        employees = self.service.get_all_employees()
        f = EmployeeFilter(request.query_params, queryset=employees)
        return f.qs

    @extend_schema(
        summary="List all employees",
        parameters=[
            OpenApiParameter("search", OpenApiTypes.STR, description="Search by first or last name"),
            OpenApiParameter("department", OpenApiTypes.STR, description="Filter by department (IT, HR, FINANCE, MARKETING, OPERATIONS, SALES)"),
            OpenApiParameter("is_active", OpenApiTypes.BOOL, description="Filter by active status"),
        ],
        responses={200: EmployeeSerializer(many=True)},
        tags=["Employees"]
    )
    def list(self, request):
        employees = self._get_filtered_employees(request)
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Retrieve an employee",
        responses={200: EmployeeSerializer},
        tags=["Employees"]
    )
    def retrieve(self, request, pk=None):
        try:
            employee = self.service.get_employee_by_id(pk)
        except ValidationError as e:
            return Response({"detail": e.messages[0]}, status=status.HTTP_404_NOT_FOUND)
        serializer = EmployeeSerializer(employee)
        return Response(serializer.data)

    @extend_schema(
        summary="Create an employee",
        request=EmployeeCreateSerializer,
        responses={201: EmployeeSerializer},
        tags=["Employees"]
    )
    def create(self, request):
        serializer = EmployeeCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            employee = self.service.create_employee(serializer.validated_data)
        except ValidationError as e:
            return Response(e.message_dict, status=status.HTTP_400_BAD_REQUEST)
        return Response(EmployeeSerializer(employee).data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Update an employee",
        request=EmployeeUpdateSerializer,
        responses={200: EmployeeSerializer},
        tags=["Employees"]
    )
    def update(self, request, pk=None):
        serializer = EmployeeUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            employee = self.service.update_employee(pk, serializer.validated_data)
        except ValidationError as e:
            return Response(e.message_dict, status=status.HTTP_400_BAD_REQUEST)
        return Response(EmployeeSerializer(employee).data)

    @extend_schema(
        summary="Partially update an employee",
        request=EmployeeUpdateSerializer,
        responses={200: EmployeeSerializer},
        tags=["Employees"]
    )
    def partial_update(self, request, pk=None):
        serializer = EmployeeUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        try:
            employee = self.service.update_employee(pk, serializer.validated_data)
        except ValidationError as e:
            return Response(e.message_dict, status=status.HTTP_400_BAD_REQUEST)
        return Response(EmployeeSerializer(employee).data)

    @extend_schema(
        summary="Soft delete an employee",
        responses={204: None},
        tags=["Employees"]
    )
    def destroy(self, request, pk=None):
        try:
            self.service.delete_employee(pk)
        except ValidationError as e:
            return Response({"detail": e.messages[0]}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @extend_schema(
        summary="Export employees to Excel or PDF",
        parameters=[
            OpenApiParameter("export_format", OpenApiTypes.STR, description="Export format: excel or pdf"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Search by name"),
            OpenApiParameter("department", OpenApiTypes.STR, description="Filter by department"),
            OpenApiParameter("is_active", OpenApiTypes.BOOL, description="Filter by active status"),
        ],
        tags=["Employees"]
    )
    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        export_format = request.query_params.get('format', 'excel').lower()
        employees = self._get_filtered_employees(request)

        if export_format == 'excel':
            return self._export_excel(employees)
        elif export_format == 'pdf':
            return self._export_pdf(employees)
        else:
            return Response(
                {"detail": "Invalid format. Use 'excel' or 'pdf'."},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _export_excel(self, employees):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ["Code", "First Name", "Last Name", "Email", "Phone", "Department", "Position", "Date Joined", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row, emp in enumerate(employees, 2):
            ws.cell(row=row, column=1, value=emp.employee_code)
            ws.cell(row=row, column=2, value=emp.first_name)
            ws.cell(row=row, column=3, value=emp.last_name)
            ws.cell(row=row, column=4, value=emp.email)
            ws.cell(row=row, column=5, value=emp.phone_number or '')
            ws.cell(row=row, column=6, value=emp.department)
            ws.cell(row=row, column=7, value=emp.position)
            ws.cell(row=row, column=8, value=str(emp.date_joined))
            ws.cell(row=row, column=9, value="Active" if emp.is_active else "Inactive")

            # Alternate row color
            if row % 2 == 0:
                fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = fill

        # Auto column width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
        return response

    def _export_pdf(self, employees):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title = Paragraph("Employee Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3 * inch))

        # Table data
        data = [["Code", "Full Name", "Email", "Department", "Position", "Status"]]
        for emp in employees:
            data.append([
                emp.employee_code,
                emp.full_name,
                emp.email,
                emp.department,
                emp.position,
                "Active" if emp.is_active else "Inactive"
            ])

        # Table style
        table = Table(data, colWidths=[1*inch, 2*inch, 2.5*inch, 1.5*inch, 2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="employees.pdf"'
        return response