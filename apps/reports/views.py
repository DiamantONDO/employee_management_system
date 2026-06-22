from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io

from apps.reports.services import ReportService


class EmployeeSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Get employee attendance summary",
        parameters=[
            OpenApiParameter("employee_id", OpenApiTypes.UUID, required=True, description="Employee UUID"),
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True, description="Start date (YYYY-MM-DD)"),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True, description="End date (YYYY-MM-DD)"),
        ],
        tags=["Reports"]
    )
    def get(self, request):
        employee_id = request.query_params.get('employee_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if not all([employee_id, start_date, end_date]):
            return Response(
                {"detail": "employee_id, start_date and end_date are required."},
                status=400
            )
        try:
            summary = ReportService().get_employee_summary(employee_id, start_date, end_date)
        except ValidationError as e:
            return Response(e.message_dict, status=404)
        return Response(summary)


class EmployeeSummaryExportView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Export employee summary to PDF or Excel",
        parameters=[
            OpenApiParameter("employee_id", OpenApiTypes.UUID, required=True),
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("export_format", OpenApiTypes.STR, description="excel or pdf"),
        ],
        tags=["Reports"]
    )
    def get(self, request):
        employee_id = request.query_params.get('employee_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export_format = request.query_params.get('format', 'pdf').lower()

        if not all([employee_id, start_date, end_date]):
            return Response({"detail": "employee_id, start_date and end_date are required."}, status=400)

        try:
            summary = ReportService().get_employee_summary(employee_id, start_date, end_date)
        except ValidationError as e:
            return Response(e.message_dict, status=404)

        if export_format == 'excel':
            return self._export_excel(summary)
        return self._export_pdf(summary)

    def _export_excel(self, summary):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employee Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        rows = [
            ("Employee", summary['employee_name']),
            ("Total Days", summary['total_days']),
            ("Present Days", summary['present_days']),
            ("Absent Days", summary['absent_days']),
            ("Attendance %", f"{summary['attendance_percentage']}%"),
        ]

        for i, (metric, value) in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=metric)
            ws.cell(row=i, column=2, value=value)
            if i % 2 == 0:
                fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
                for col in range(1, 3):
                    ws.cell(row=i, column=col).fill = fill

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employee_summary_{summary["employee_name"]}.xlsx"'
        return response

    def _export_pdf(self, summary):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Employee Attendance Summary", styles['Title']))
        elements.append(Paragraph(f"Employee: {summary['employee_name']}", styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        data = [
            ["Metric", "Value"],
            ["Total Days", str(summary['total_days'])],
            ["Present Days", str(summary['present_days'])],
            ["Absent Days", str(summary['absent_days'])],
            ["Attendance %", f"{summary['attendance_percentage']}%"],
        ]

        table = Table(data, colWidths=[3*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="employee_summary_{summary["employee_name"]}.pdf"'
        return response


class DepartmentSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Get department attendance summary",
        parameters=[
            OpenApiParameter("department", OpenApiTypes.STR, required=True),
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
        ],
        tags=["Reports"]
    )
    def get(self, request):
        department = request.query_params.get('department')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if not all([department, start_date, end_date]):
            return Response(
                {"detail": "department, start_date and end_date are required."},
                status=400
            )
        summary = ReportService().get_department_summary(department, start_date, end_date)
        return Response(summary)


class DepartmentSummaryExportView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Export department summary to PDF or Excel",
        parameters=[
            OpenApiParameter("department", OpenApiTypes.STR, required=True),
            OpenApiParameter("start_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("end_date", OpenApiTypes.DATE, required=True),
            OpenApiParameter("export_format", OpenApiTypes.STR, description="excel or pdf"),
        ],
        tags=["Reports"]
    )
    def get(self, request):
        department = request.query_params.get('department')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        export_format = request.query_params.get('export_format', 'pdf').lower()

        if not all([department, start_date, end_date]):
            return Response({"detail": "department, start_date and end_date are required."}, status=400)

        summary = ReportService().get_department_summary(department, start_date, end_date)

        if export_format == 'excel':
            return self._export_excel(summary)
        return self._export_pdf(summary)

    def _export_excel(self, summary):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Department Summary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        for col, header in enumerate(["Metric", "Value"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        rows = [
            ("Department", summary['department']),
            ("Total Employees", summary['total_employees']),
            ("Total Days", summary['total_days']),
            ("Present Days", summary['present_days']),
            ("Absent Days", summary['absent_days']),
            ("Attendance %", f"{summary['attendance_percentage']}%"),
        ]

        for i, (metric, value) in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=metric)
            ws.cell(row=i, column=2, value=value)
            if i % 2 == 0:
                fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
                for col in range(1, 3):
                    ws.cell(row=i, column=col).fill = fill

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="department_summary_{summary["department"]}.xlsx"'
        return response

    def _export_pdf(self, summary):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Department Attendance Summary", styles['Title']))
        elements.append(Paragraph(f"Department: {summary['department']}", styles['Normal']))
        elements.append(Spacer(1, 0.3 * inch))

        data = [
            ["Metric", "Value"],
            ["Total Employees", str(summary['total_employees'])],
            ["Total Days", str(summary['total_days'])],
            ["Present Days", str(summary['present_days'])],
            ["Absent Days", str(summary['absent_days'])],
            ["Attendance %", f"{summary['attendance_percentage']}%"],
        ]

        table = Table(data, colWidths=[3*inch, 3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="department_summary_{summary["department"]}.pdf"'
        return response