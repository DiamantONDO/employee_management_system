from rest_framework import status
from rest_framework.viewsets import ViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.core.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiParameter
from drf_spectacular.types import OpenApiTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
import io

from apps.attendance.services import AttendanceService
from apps.attendance.serializers import (
    AttendanceSerializer,
    AttendanceCreateSerializer,
    AttendanceUpdateSerializer
)
from apps.attendance.filters import AttendanceFilter


class AttendanceViewSet(ViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = AttendanceFilter

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.service = AttendanceService()

    def _get_filtered_records(self, request):
        records = self.service.get_all_attendance()
        f = AttendanceFilter(request.query_params, queryset=records)
        return f.qs

    @extend_schema(
        summary="List all attendance records",
        parameters=[
            OpenApiParameter("status", OpenApiTypes.STR, description="Filter by status (PRESENT, ABSENT)"),
            OpenApiParameter("date", OpenApiTypes.DATE, description="Filter by specific date (YYYY-MM-DD)"),
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Filter from this date (YYYY-MM-DD)"),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Filter to this date (YYYY-MM-DD)"),
            OpenApiParameter("employee_id", OpenApiTypes.UUID, description="Filter by employee UUID"),
        ],
        responses={200: AttendanceSerializer(many=True)},
        tags=["Attendance"]
    )
    def list(self, request):
        records = self._get_filtered_records(request)
        serializer = AttendanceSerializer(records, many=True)
        return Response(serializer.data)

    @extend_schema(
        summary="Retrieve an attendance record",
        responses={200: AttendanceSerializer},
        tags=["Attendance"]
    )
    def retrieve(self, request, pk=None):
        try:
            attendance = self.service.get_attendance_by_id(pk)
        except ValidationError as e:
            return Response({"detail": e.messages[0]}, status=status.HTTP_404_NOT_FOUND)
        serializer = AttendanceSerializer(attendance)
        return Response(serializer.data)

    @extend_schema(
        summary="Create an attendance record",
        request=AttendanceCreateSerializer,
        responses={201: AttendanceSerializer},
        tags=["Attendance"]
    )
    def create(self, request):
        serializer = AttendanceCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            attendance = self.service.create_attendance(serializer.validated_data)
        except ValidationError as e:
            return Response(e.message_dict, status=status.HTTP_400_BAD_REQUEST)
        return Response(AttendanceSerializer(attendance).data, status=status.HTTP_201_CREATED)

    @extend_schema(
        summary="Update an attendance record",
        request=AttendanceUpdateSerializer,
        responses={200: AttendanceSerializer},
        tags=["Attendance"]
    )
    def update(self, request, pk=None):
        serializer = AttendanceUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            attendance = self.service.update_attendance(pk, serializer.validated_data)
        except ValidationError as e:
            return Response(e.message_dict, status=status.HTTP_400_BAD_REQUEST)
        return Response(AttendanceSerializer(attendance).data)

    @extend_schema(
        summary="Partially update an attendance record",
        request=AttendanceUpdateSerializer,
        responses={200: AttendanceSerializer},
        tags=["Attendance"]
    )
    def partial_update(self, request, pk=None):
        serializer = AttendanceUpdateSerializer(data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        try:
            attendance = self.service.update_attendance(pk, serializer.validated_data)
        except ValidationError as e:
            return Response(e.message_dict, status=status.HTTP_400_BAD_REQUEST)
        return Response(AttendanceSerializer(attendance).data)

    @extend_schema(
        summary="Delete an attendance record",
        responses={204: None},
        tags=["Attendance"]
    )
    def destroy(self, request, pk=None):
        try:
            self.service.delete_attendance(pk)
        except ValidationError as e:
            return Response({"detail": e.messages[0]}, status=status.HTTP_404_NOT_FOUND)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @extend_schema(
        summary="Export attendance to Excel or PDF",
        parameters=[
            OpenApiParameter("export_format", OpenApiTypes.STR, description="Export format: excel or pdf"),
            OpenApiParameter("status", OpenApiTypes.STR, description="Filter by status"),
            OpenApiParameter("date", OpenApiTypes.DATE, description="Filter by date"),
            OpenApiParameter("start_date", OpenApiTypes.DATE, description="Filter from date"),
            OpenApiParameter("end_date", OpenApiTypes.DATE, description="Filter to date"),
            OpenApiParameter("employee_id", OpenApiTypes.UUID, description="Filter by employee"),
        ],
        tags=["Attendance"]
    )
    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        export_format = request.query_params.get('export_format', 'excel').lower()
        records = self._get_filtered_records(request)

        if export_format == 'excel':
            return self._export_excel(records)
        elif export_format == 'pdf':
            return self._export_pdf(records)
        else:
            return Response(
                {"detail": "Invalid format. Use 'excel' or 'pdf'."},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _export_excel(self, records):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Employee", "Date", "Check In", "Check Out", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.employee.full_name)
            ws.cell(row=row, column=2, value=str(record.attendance_date))
            ws.cell(row=row, column=3, value=str(record.check_in_time or ''))
            ws.cell(row=row, column=4, value=str(record.check_out_time or ''))
            ws.cell(row=row, column=5, value=record.status)

            if row % 2 == 0:
                fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = fill

        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="attendance.xlsx"'
        return response

    def _export_pdf(self, records):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        title = Paragraph("Attendance Report", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3 * inch))

        data = [["Employee", "Date", "Check In", "Check Out", "Status"]]
        for record in records:
            data.append([
                record.employee.full_name,
                str(record.attendance_date),
                str(record.check_in_time or ''),
                str(record.check_out_time or ''),
                record.status,
            ])

        table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="attendance.pdf"'
        return response